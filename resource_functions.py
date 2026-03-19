import os
import sys
from openpyxl import load_workbook


def get_resource(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def get_sheet_names(file):
	wb = load_workbook(file, read_only=True)
	sheet_names = wb.sheetnames
	wb.close()

	return sheet_names