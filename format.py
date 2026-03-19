import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side

def get_resource(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        # In PyInstaller 6+ onedir mode, this points to the '_internal' folder automatically
        base_path = sys._MEIPASS
    except Exception:
        # In development, use the current directory or the script's directory
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# Column letters and widths
columns_size = {'A':11.75, 'All':10.25, 'J':6.25, 'K':9.89, 'M':11.25, 'N':11.04, 'O':14.89, 'P':13.20}

def format(sheet):
	total_format = NamedStyle(name="total_format")
	total_format.font = Font(bold=True)

	wb = load_workbook(get_resource('Total Time Worked.xlsx'))
	ws = wb[sheet]

	# Register name style
	if "total_format" not in wb.named_styles:
		wb.add_named_style(total_format)

	# Apply styles for attendants
	if sheet == 'Attendants':
		# SET BOLD
		for row in range(1, 3):
			for col in range(1,16):
				ws.cell(row, col).style = 'total_format'

		for row in range(16, 18):
			for col in range(1,21):
				ws.cell(row, col).style = 'total_format'

		# SET COLUMN WIDTH
		# Column A 
		ws.column_dimensions['A'].width = columns_size['A']

		# Column B:H
		for col_num in range(2, 9):  
			col = get_column_letter(col_num)
			ws.column_dimensions[col].width = columns_size['All']

		# Column J:P
		ws.column_dimensions['J'].width = columns_size['J']
		ws.column_dimensions['K'].width = columns_size['K']
		ws.column_dimensions['M'].width = columns_size['M']
		ws.column_dimensions['N'].width = columns_size['N']
		ws.column_dimensions['O'].width = columns_size['O']
		ws.column_dimensions['P'].width = columns_size['P']

		# SET COLUMN ALIGNMENT B:P
		for row in range(1, ws.max_row + 1):
			for col in range(2,17):
				ws.cell(row, col).alignment = Alignment(horizontal='center')
		
		# Set column M back to left 
		for cell in ws['M']:
			cell.alignment = Alignment(horizontal='left')

	# Apply styles for cashiers and bakers
	else:
		# SET BOLD
		for row in range(1, 3):
			for col in range(1,16):
				ws.cell(row, col).style = 'total_format'
		
		for row in range(8, 10):
			for col in range(1,16):
				ws.cell(row, col).style = 'total_format'

		for row in range(16, 18):
			for col in range(1,21):
				ws.cell(row, col).style = 'total_format'

		for row in range(23, 25):
			for col in range(1,21):
				ws.cell(row, col).style = 'total_format'

		# SET COLUMN WIDTH
		# Column A 
		ws.column_dimensions['A'].width = columns_size['A']

		# Column B:H
		for col_num in range(2, 9):  
			col = get_column_letter(col_num)
			ws.column_dimensions[col].width = columns_size['All']

		# Column J:O
		ws.column_dimensions['J'].width = columns_size['J']
		ws.column_dimensions['K'].width = columns_size['K']
		ws.column_dimensions['M'].width = columns_size['M']
		ws.column_dimensions['N'].width = columns_size['N']
		ws.column_dimensions['O'].width = columns_size['O']
		ws.column_dimensions['P'].width = columns_size['P']

		# SET COLUMN ALIGNMENT B:P
		for row in range(1, ws.max_row + 1):
			for col in range(2,17):
				ws.cell(row, col).alignment = Alignment(horizontal='center')
		
		# Set column M back to left 		
		for cell in ws['M']:
			cell.alignment = Alignment(horizontal='left')

	wb.save(get_resource('Total Time Worked.xlsx'))
	wb.close()