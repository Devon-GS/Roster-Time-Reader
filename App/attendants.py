# import pandas as pd
# import re

# # Notes
# # before dash [0-9]+(?=.*\-)
# # after dash \-(.*)

# def attendant_one(selected):
#     week_one_list = []

#     # Read file and get sheets
#     file = '../Attendant_Carwash_Roster.xlsx'
#     file_sheets = pd.ExcelFile(file).sheet_names
#     # Get Columns
#     columns = ['idx','ATTENDANTS', 'THURS', 'FRI', 'SAT', 'SUN', 'MON', 'TUE', 'WED']
#     # Get last sheet
#     sheet = file_sheets[selected]
#     # Read in file last sheet
#     data = pd.read_excel(file, sheet_name=sheet, index_col=0, header=1, usecols=columns)

#     # Get week one from excel sheet
#     week_one_data = data.loc[0:14]
#     week_one = []
#     for x in week_one_data.to_numpy():
#         if str(x[0]) != 'nan':
#             week_one.append(x)
	
#     # FUNCTIONS
#     # Find numbers before dash and after dash
#     def first(weekday):
#         first = float(re.findall('[0-9]+(?=.*\-)', weekday)[0])
#         return first

#     def second(weekday):
#         second = float(re.findall('\-(.*)', weekday)[0])
#         return second

#     # Workings
#     for week in week_one:
#         name = week[0]
#         thur = week[1]
#         fri = week[2]
#         sat = week[3]
#         sun = week[4]
#         mon = week[5]
#         tue = week[6]
#         wed = week[7]
#         sune = 0.0
#         mone = 0.0

#         if thur == 'AF':
#             thur = 0
#         elif second(thur) == 6.3:
#             thur = (24 - first(thur)) + 6.5
#         elif first(thur) >= 18:
#             thur = (24 - first(thur)) + second(thur)
#         elif first(thur) - second(thur) < 0:
#             thur = (first(thur) - second(thur)) * -1
#         else:
#             thur = first(thur) - second(thur)
		
#         if fri == 'AF':
#             fri = 0
#         elif second(fri) == 6.3:
#             fri = (24 - first(fri)) + 6.5
#         elif first(fri) >= 18:
#             fri = (24 - first(fri)) + second(fri)
#         elif first(fri) - second(fri) < 0:
#             fri = (first(fri) - second(fri)) * -1
#         else:
#             fri = first(fri) - second(fri)

#         if sat == 'AF':
#             sat = 0
#         elif sat == '18-6' or sat == '18-7':
#             sune += second(sat)
#             sat = (24 - first(sat))
#         elif sat == "18-6.30":
#             sune += 6.5
#             sat = (24 - first(sat))
#         elif sat == '6.30-18':
#             sat = second(sat) - 6.5
#         else:
#             sat = (first(sat) - second(sat)) * -1

#         if sun == 'AF':
#             sun = 0
#         elif sun == '18-6' or sun == '18-7':
#             mone += second(sun)
#             sun = (24 - first(sun))
#         elif sun == '18-6.30':
#             mone += 6.5
#             sun = (24 - first(sun))
#         elif sun == '6.30-18':
#             sun = second(sun) - 6.5
#         else:
#             sun = (first(sun) - second(sun)) * -1

#         if mon == 'AF':
#             mon = 0
#         elif second(mon) == 6.3:
#             mon = (24 - first(mon)) + 6.5
#         elif first(mon) >= 18:
#             mon = (24 - first(mon)) + second(mon)
#         elif first(mon) - second(mon) < 0:
#             mon = (first(mon) - second(mon)) * -1
#         else:
#             mon = first(mon) - second(mon)

#         if tue == 'AF':
#             tue = 0
#         elif second(tue) == 6.3:
#             tue = (24 - first(tue)) + 6.5
#         elif first(tue) >= 18:
#             tue = (24 - first(tue)) + second(tue)
#         elif first(tue) - second(tue) < 0:
#             tue = (first(tue) - second(tue)) * -1
#         else:
#             tue = first(tue) - second(tue)

#         if wed == 'AF':
#             wed = 0
#         elif second(wed) == 6.3:
#             wed = (24 - first(wed)) + 6.5
#         elif first(wed) >= 18:
#             wed = (24 - first(wed)) + second(wed)
#         elif first(wed) - second(wed) < 0:
#             wed = (first(wed) - second(wed)) * -1
#         else:
#             wed = first(wed) - second(wed)

#         total_hours_week_one = thur + fri + sat + mon + mone + tue + wed
#         total_sun_week_one = sun + sune
#         week_one_list.append([name , total_hours_week_one , total_sun_week_one])
#     return week_one_list

# def attendant_two(selected):
#     week_two_list = []

#     # Read file and get sheets
#     file = '../Attendant_Carwash_Roster.xlsx'
#     file_sheets = pd.ExcelFile(file).sheet_names
#     # Get Columns
#     columns = ['idx','ATTENDANTS', 'THURS', 'FRI', 'SAT', 'SUN', 'MON', 'TUE', 'WED']
#     # Get last sheet
#     sheet = file_sheets[selected]
#     # Read in file last sheet
#     data = pd.read_excel(file, sheet_name=sheet, index_col=0, header=1, usecols=columns)

#     # Get week two from excel sheet
#     week_two_data = data.loc[30:44]
	
#     week_two = []
#     for x in week_two_data.to_numpy():
#         if str(x[0]) != 'nan':
#             week_two.append(x)

#     # FUNCTIONS
#     # Find numbers before dash and after dash
#     def first(weekday):
#         first = float(re.findall('[0-9]+(?=.*\-)', weekday)[0])
#         return first

#     def second(weekday):
#         second = float(re.findall('\-(.*)', weekday)[0])
#         return second

#     for week in week_two:
#         name = week[0]
#         thur = week[1]
#         fri = week[2]
#         sat = week[3]
#         sun = week[4]
#         mon = week[5]
#         tue = week[6]
#         wed = week[7]
#         sune = 0.0
#         mone = 0.0

#         if thur == 'AF':
#             thur = 0
#         elif second(thur) == 6.3:
#             thur = (24 - first(thur)) + 6.5
#         elif first(thur) >= 18:
#             thur = (24 - first(thur)) + second(thur)
#         elif first(thur) - second(thur) < 0:
#             thur = (first(thur) - second(thur)) * -1
#         else:
#             thur = first(thur) - second(thur)
		
#         if fri == 'AF':
#             fri = 0
#         elif second(fri) == 6.3:
#             fri = (24 - first(fri)) + 6.5
#         elif first(fri) >= 18:
#             fri = (24 - first(fri)) + second(fri)
#         elif first(fri) - second(fri) < 0:
#             fri = (first(fri) - second(fri)) * -1
#         else:
#             fri = first(fri) - second(fri)

#         if sat == 'AF':
#             sat = 0
#         elif sat == '18-6' or sat == '18-7':
#             sune += second(sat)
#             sat = (24 - first(sat))
#         elif sat == "18-6.30":
#             sune += 6.5
#             sat = (24 - first(sat))
#         elif sat == '6.30-18':
#             sat = second(sat) - 6.5
#         else:
#             sat = (first(sat) - second(sat)) * -1

#         if sun == 'AF':
#             sun = 0
#         elif sun == '18-6' or sun == '18-7':
#             mone += second(sun)
#             sun = (24 - first(sun))
#         elif sun == '18-6.30':
#             mone += 6.5
#             sun = (24 - first(sun))
#         elif sun == '6.30-18':
#             sun = second(sun) - 6.5
#         else:
#             sun = (first(sun) - second(sun)) * -1

#         if mon == 'AF':
#             mon = 0
#         elif second(mon) == 6.3:
#             mon = (24 - first(mon)) + 6.5
#         elif first(mon) >= 18:
#             mon = (24 - first(mon)) + second(mon)
#         elif first(mon) - second(mon) < 0:
#             mon = (first(mon) - second(mon)) * -1
#         else:
#             mon = first(mon) - second(mon)

#         if tue == 'AF':
#             tue = 0
#         elif second(tue) == 6.3:
#             tue = (24 - first(tue)) + 6.5
#         elif first(tue) >= 18:
#             tue = (24 - first(tue)) + second(tue)
#         elif first(tue) - second(tue) < 0:
#             tue = (first(tue) - second(tue)) * -1
#         else:
#             tue = first(tue) - second(tue)

#         if wed == 'AF':
#             wed = 0
#         elif second(wed) == 6.3:
#             wed = (24 - first(wed)) + 6.5
#         elif first(wed) >= 18:
#             wed = (24 - first(wed)) + second(wed)
#         elif first(wed) - second(wed) < 0:
#             wed = (first(wed) - second(wed)) * -1
#         else:
#             wed = first(wed) - second(wed)

#         total_hours_week_two = thur + fri + sat + mon + mone + tue + wed
#         total_sun_week_two = sun + sune
#         week_two_list.append([name , total_hours_week_two , total_sun_week_two])
#     return week_two_list


# #######################################################################################################
# #######################################################################################################
# #######################################################################################################
# #######################################################################################################

# ==============================================================================
# ATTENDENTS WEEK 1 - ROSTER TIMES TO EXCEL
# ==============================================================================
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import re
import sqlite3

# ==============================================================================
# IMPORT ROSTER TIMES AND DATES AND BADGE WEEK 1 (ATTENDENTS)
# ==============================================================================
file = '../Attendant_Carwash_Roster.xlsx'

# Get Columns
columns = ['idx','ATTENDANTS', 'THURS', 'FRI', 'SAT', 'SUN', 'MON', 'TUE', 'WED']

# Get Times
data = pd.read_excel(file, index_col=0, header=1, usecols=columns)
data = data.fillna(0)

# Get Dates
data_date = pd.read_excel(file, header=None, nrows=2, usecols='C:I')
data_date_ex = data_date.loc[0]

weekone_dates = {}
for x in data_date_ex:
	weekone_dates[x.strftime("%A")] = x.date().strftime("%d/%m/%y")

# Get week one from excel roster sheet
week_one_data = data.loc[0:14]
week_one = []
for x in week_one_data.to_numpy().tolist():
	if str(x[0]) != 'nan':
		if x[0] != 0:
			week_one.append(x)

# ==============================================================================
# CREATE DATABASE SQLITE WEEK 1
# ==============================================================================
def db_init():
	# Connect to database
	con = sqlite3.connect("database/time_sheet.db")
	c = con.cursor()
	# Create table
	c.execute(
		"""CREATE TABLE IF NOT EXISTS rosterAttWeekOne (
			name TEXT,
			badge TEXT,
			thur TEXT,
			fri TEXT,
			sat TEXT,
			sun TEXT,
			mon TEXT,
			tue TEXT,
			wed TEXT
			)"""
	)

	# Add week one data to table
	week1 = """INSERT INTO rosterAttWeekOne (
			name,
			badge,
			thur,
			fri,
			sat,
			sun,
			mon,
			tue,
			wed
			)
			VALUES (?, ?, ?, ?, ? ,? ,?, ?, ?)"""

	c.execute(week1, ("Date", "999", 0, 0, 0, 0, 0, 0, 0))

	for week in week_one:
		x = (week[0], 0, week[1], week[2], week[3], week[4], week[5], week[6], week[7])
		c.execute(week1, (x))

		con.commit()
	con.close()

def db_update_dates():
	# Update table with roster dates
	con = sqlite3.connect("database/time_sheet.db")
	c = con.cursor()

	query = """UPDATE rosterAttWeekOne SET
			thur = ?,
			fri = ?,
			sat = ?,
			sun = ?,
			mon = ?,
			tue = ?,
			wed = ?
			WHERE badge = ?
			"""
	thursday = weekone_dates["Thursday"]
	friday = weekone_dates["Friday"]
	saturday = weekone_dates["Saturday"]
	sunday = weekone_dates["Sunday"]
	monday = weekone_dates["Monday"]
	tuesday = weekone_dates["Tuesday"]
	wednesday = weekone_dates["Wednesday"]

	c.execute(
		query, (thursday, friday, saturday, sunday, monday, tuesday, wednesday, 999)
	)
	con.commit()

# db_init()
# db_update_dates()

# ==============================================================================
# ATTENDANTS WEEK 1 - To Excel
# ==============================================================================

def attendant_one_excel():
	# Grab data from table for excel workbook
	con = sqlite3.connect("database/time_sheet.db")
	c = con.cursor()

	c.execute("SELECT name FROM rosterAttWeekOne")
	name_records = c.fetchall()

	week_one_data = []

	for record in name_records:
		# Grab data from database using name of person
		c.execute("SELECT * FROM rosterAttWeekOne where name=?", (record[0],))
		r = c.fetchall()
		if record[0] != 'Date':
			week_one_data.append(r)

	con.close()

	# FUNCTIONS
	# Find numbers before dash and after dash
	def first(weekday):
		# print(weekday)
		first = float(re.findall('[0-9]+(?=.*\-)', weekday)[0])
		return first

	def second(weekday):
		second = float(re.findall('\-(.*)', weekday)[0])
		return second
	
	# ##############################################
	# BUILD OUT EXCEL FILE
	# ##############################################

	# Start workbook
	wb = Workbook()
	wb.active.title = 'Attendents'
	ws = wb['Attendents']

	# Get days of week
	thursday = weekone_dates["Thursday"]
	friday = weekone_dates["Friday"]
	saturday = weekone_dates["Saturday"]
	sunday = weekone_dates["Sunday"]
	monday = weekone_dates["Monday"]
	tuesday = weekone_dates["Tuesday"]
	wednesday = weekone_dates["Wednesday"]
	# thursday_t = datetime.strptime(weekone_dates["Wednesday"], "%d/%m/%y") + timedelta(days=1)
	# thursday_s = thursday_t.strftime("%d/%m/%y")

	# Headings
	ws['A1'] = 'Week 1'
	ws['B1'] = thursday
	ws['C1'] = friday
	ws['D1'] = saturday
	ws['E1'] = sunday
	ws['F1'] = monday
	ws['G1'] = tuesday
	ws['H1'] = wednesday

	ws['A2'] = 'ATTENDENTS'
	ws['B2'] = 'THURS'
	ws['C2'] = 'FRI'
	ws['D2'] = 'SAT'
	ws['E2'] = 'SUN'
	ws['F2'] = 'MON'
	ws['G2'] = 'TUE'
	ws['H2'] = 'WED'

	# Row and column start
	row = 3
	col = 0
	i = 0

	# Extract data to excel
	for r in week_one_data[1:]:
		# print(r)
		name = r[0][0]
		thur = r[0][2]
		fri = r[0][3]
		sat = r[0][4]
		sun = r[0][5]
		mon = r[0][6]
		tue = r[0][7]
		wed = r[0][8]

		ws[f'A{row + i}'] = name

		if thur == 'AF':
			th = 0
			ws[f'B{row + i}'] = f'{thur} ({int(th)})'
		elif first(thur) >= 18:
			th = (24 - first(thur)) + second(thur)
			ws[f'B{row + i}'] = f'{thur} ({int(th)})'
		elif first(thur) - second(thur) < 0:
			th = (first(thur) - second(thur)) * -1
			ws[f'B{row + i}'] = f'{thur} ({int(th)})'
		else:
			th = first(thur) - second(thur)
			ws[f'B{row + i}'] = f'{thur} ({int(th)})'
		
		# if fri == 'AF':
		# 	fr = 0
		# elif first(fri) >= 18:
		# 	fr = (24 - first(fri)) + second(fri)
		# elif first(fri) - second(fri) < 0:
		# 	fr = (first(fri) - second(fri)) * -1
		# else:
		# 	fr = first(fri) - second(fri)

		# if sat == 'AF':
		# 	sa = 0
		# elif sat == '18-6' or sat == '18-7':
		# 	sun_n = second(sat)
		# 	sa = (24 - first(sat))
		# else:
		# 	sa = (first(sat) - second(sat)) * -1

		# if sun == 'AF':
		# 	su = 0
		# elif sun == '18-6' or sun == '18-7':
		# 	mon_n = second(sun)
		# 	su = (24 - first(sun))
		# else:
		# 	su = (first(sun) - second(sun)) * -1

		# if mon == 'AF':
		# 	mo = 0
		# elif first(mon) >= 18:
		# 	mo = (24 - first(mon)) + second(mon)
		# elif first(mon) - second(mon) < 0:
		# 	mo = (first(mon) - second(mon)) * -1
		# else:
		# 	mo = first(mon) - second(mon)

		# if tue == 'AF':
		# 	tu = 0
		# elif first(tue) >= 18:
		# 	tu = (24 - first(tue)) + second(tue)
		# elif first(tue) - second(tue) < 0:
		# 	tu = (first(tue) - second(tue)) * -1
		# else:
		# 	tu = first(tue) - second(tue)

		# if wed == 'AF':
		# 	we = 0
		# elif first(wed) >= 18:
		# 	we = (24 - first(wed)) + second(wed)
		# elif first(wed) - second(wed) < 0:
		# 	we = (first(wed) - second(wed)) * -1
		# else:
		# 	we = first(wed) - second(wed)

		i += 1

	# TOTAL HOURS WEEKDAY AND SUNDAY AT END 
		
	# Close workbook
	wb.save("Total Time Worked.xlsx")
	wb.close()



attendant_one_excel()