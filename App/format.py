from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side

# Column letters and widths
columns_size = {'A':11.75, 'All':10.25, 'J':6.25, 'K':9.89, 'M':11.25, 'N':11.04, 'O':14.89}

def format(sheet):
	wb = load_workbook('_internal/Total Time Worked.xlsx')
	ws = wb[sheet]

	total_format = NamedStyle(name="total_format")
	total_format.font = Font(bold=True)

	if sheet == 'Attendants':
		# SET BOLD
		for row in range(1, 3):
			for col in range(2,16):
				ws.cell(row, col).style = total_format

		for row in range(16, 18):
			for col in range(2,21):
				ws.cell(row, col).style = 'total_format'

		# SET COLUMN WIDTH
		# Column A 
		ws.column_dimensions['A'].width = columns_size['A']

		# Column B:H
		for col_num in range(2, 9):  
			col = get_column_letter(col_num)
			ws.column_dimensions[col].width = columns_size['All']

		# Column J:O
		ws.column_dimensions['J'].width = columns_size['J']
		ws.column_dimensions['K'].width = columns_size['K']
		ws.column_dimensions['M'].width = columns_size['M']
		ws.column_dimensions['N'].width = columns_size['N']
		ws.column_dimensions['O'].width = columns_size['O']

		# SET COLUMN ALIGNMENT B:H
		for row in range(1, ws.max_row + 1):
			for col in range(2,16):
				ws.cell(row, col).alignment = Alignment(horizontal='center')
				
		for cell in ws['M']:
			cell.alignment = Alignment(horizontal='left')
	else:
		# SET BOLD
		for row in range(1, 3):
			for col in range(2,16):
				ws.cell(row, col).style = 'total_format'
		
		for row in range(8, 10):
			for col in range(2,16):
				ws.cell(row, col).style = 'total_format'

		for row in range(16, 18):
			for col in range(2,21):
				ws.cell(row, col).style = 'total_format'

		for row in range(23, 25):
			for col in range(2,21):
				ws.cell(row, col).style = 'total_format'

		# SET COLUMN WIDTH
		# Column A 
		ws.column_dimensions['A'].width = columns_size['A']

		# Column B:H
		for col_num in range(2, 9):  
			col = get_column_letter(col_num)
			ws.column_dimensions[col].width = columns_size['All']

		# Column J:O
		ws.column_dimensions['J'].width = columns_size['J']
		ws.column_dimensions['K'].width = columns_size['K']
		ws.column_dimensions['M'].width = columns_size['M']
		ws.column_dimensions['N'].width = columns_size['N']
		ws.column_dimensions['O'].width = columns_size['O']

		# SET COLUMN ALIGNMENT B:H
		for row in range(1, ws.max_row + 1):
			for col in range(2,16):
				ws.cell(row, col).alignment = Alignment(horizontal='center')
				
		for cell in ws['M']:
			cell.alignment = Alignment(horizontal='left')

	wb.save('_internal/Total Time Worked.xlsx')
	wb.close()