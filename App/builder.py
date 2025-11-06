from openpyxl import Workbook
from openpyxl import load_workbook
import os
import re


# ==============================================================================
# Build Excel
# ==============================================================================

# Track total hours
total_hours = {}

def build_excel(roster, week_data, week_dates, week): 
	# FUNCTIONS
	# Find numbers before dash and after dash
	def first(weekday):
		first = float(re.findall('[0-9]+(?=.*\-)', weekday)[0])
		return first

	def second(weekday):
		second = float(re.findall('\-(.*)', weekday)[0])
		return second
	
	# Check if Total time excel exists
	if not os.path.exists('Total Time Worked.xlsx'):
		wb = Workbook()
		wb.active.title = "Attendants"
		wb.create_sheet("Cashiers")
		wb.save('Total Time Worked.xlsx')

	# ##############################################
	# BUILD OUT EXCEL FILE
	# ##############################################

	# Start workbook or continue
	wb = load_workbook("Total Time Worked.xlsx")
	if roster == 'Bakers':
		ws = wb['Cashiers']
	else:
		ws = wb[roster]
		
	# Get dates of week 
	if week == 1:
		thursday = week_dates["Thursday"]
		friday = week_dates["Friday"]
		saturday = week_dates["Saturday"]
		sunday = week_dates["Sunday"]
		monday = week_dates["Monday"]
		tuesday = week_dates["Tuesday"]
		wednesday = week_dates["Wednesday"]
	else:
		thursday = week_dates["Thursday"]
		friday = week_dates["Friday"]
		saturday = week_dates["Saturday"]
		sunday = week_dates["Sunday"]
		monday = week_dates["Monday"]
		tuesday = week_dates["Tuesday"]
		wednesday = week_dates["Wednesday"]

	# Headings
	if week == 1:
		if roster == 'Bakers':
			hd_row = 8
			shd_row = 9
		else:
			hd_row = 1
			shd_row = 2
	else:
		if roster == 'Bakers':
			hd_row = 23
			shd_row = 24
		else:
			hd_row = 16
			shd_row = 17

	ws[f'A{hd_row}'] = 'Week 1'
	ws[f'B{hd_row}'] = thursday
	ws[f'C{hd_row}'] = friday
	ws[f'D{hd_row}'] = saturday
	ws[f'E{hd_row}'] = sunday
	ws[f'F{hd_row}'] = monday
	ws[f'G{hd_row}'] = tuesday
	ws[f'H{hd_row}'] = wednesday

	# ws[f'A{shd_row}'] = 'ATTENDENTS'
	ws[f'B{shd_row}'] = 'THURS'
	ws[f'C{shd_row}'] = 'FRI'
	ws[f'D{shd_row}'] = 'SAT'
	ws[f'E{shd_row}'] = 'SUN'
	ws[f'F{shd_row}'] = 'MON'
	ws[f'G{shd_row}'] = 'TUE'
	ws[f'H{shd_row}'] = 'WED'

	ws[f'J{shd_row}'] = 'Hours'
	ws[f'K{shd_row}'] = 'Sun Hours'

	ws['M17'] = 'Name'
	ws['N17'] = 'Total Hours'
	ws['O17'] = 'Total Sun Hours'

	# Row and column start
	if week == 1:
		if roster == 'Bakers':
			row = 10
		else:
			row = 3
	else:
		if roster == 'Bakers':
			row = 25
		else:
			row = 18	

	# Extract data to excel
	i = 0

	for r in week_data:
		# print(r)
		name = r[0][0]
		thur = r[0][2]
		fri = r[0][3]
		sat = r[0][4]
		sun = r[0][5]
		mon = r[0][6]
		tue = r[0][7]
		wed = r[0][8]

		ws[f'A{row + i}'] = name

		if 'cas' in thur.lower():
			th = 0
			ws[f'B{row + i}'] = thur		
		elif thur == 'AF':
			th = 0
			ws[f'B{row + i}'] = f'{thur} | ({int(th)})'
		elif first(thur) >= 18:
			th = (24 - first(thur)) + second(thur)
			ws[f'B{row + i}'] = f'{thur} | ({int(th)})'
		elif first(thur) - second(thur) < 0:
			th = (first(thur) - second(thur)) * -1
			ws[f'B{row + i}'] = f'{thur} | ({int(th)})'
		else:
			th = first(thur) - second(thur)
			ws[f'B{row + i}'] = f'{thur} | ({int(th)})'
		
		if 'cas' in fri.lower():
			fr = 0
			ws[f'C{row + i}'] = fri
		elif fri == 'AF':
			fr = 0
			ws[f'C{row + i}'] = f'{fri} | ({int(fr)})'
		elif first(fri) >= 18:
			fr = (24 - first(fri)) + second(fri)
			ws[f'C{row + i}'] = f'{fri} | ({int(fr)})'
		elif first(fri) - second(fri) < 0:
			fr = (first(fri) - second(fri)) * -1
			ws[f'C{row + i}'] = f'{fri} | ({int(fr)})'
		else:
			fr = first(fri) - second(fri)
			ws[f'C{row + i}'] = f'{fri} | ({int(fr)})'

		# Certain employees work some sunday hours that have a different pay structure var_n
		global sun_wo
		if 'cas' in sat.lower():
			sa = 0
			sun_wo = 0
			ws[f'D{row + i}'] = sat
		elif sat == 'AF':
			sa = 0
			sun_wo = 0
			ws[f'D{row + i}'] = f'{sat} | ({int(sa)})'
		elif sat == '18-6' or sat == '18-7':
			sun_wo = second(sat)
			sa = (24 - first(sat))
			ws[f'D{row + i}'] = f'{sat} | ({int(sa)})'
		else:
			sun_wo = 0
			sa = (first(sat) - second(sat)) * -1
			ws[f'D{row + i}'] = f'{sat} | ({int(sa)})'

		global mon_wo
		if 'cas' in sun.lower():
			mon_wo = 0
			su = 0
			ws[f'E{row + i}'] = sun
		elif sun == 'AF':
			mon_wo = 0
			su = 0
			ws[f'E{row + i}'] = f'{sun} | ({int(su) + int(sun_wo)})'
		elif sun == '18-6' or sun == '18-7':
			mon_wo = second(sun)
			su = (24 - first(sun))
			ws[f'E{row + i}'] = f'{sun} | ({int(su) + int(sun_wo)})'
		else:
			mon_wo = 0
			su = (first(sun) - second(sun)) * -1
			ws[f'E{row + i}'] = f'{sun} | ({int(su) + int(sun_wo)})'

		if 'cas' in mon.lower():
			mo = 0
			ws[f'F{row + i}'] = mon
		elif mon == 'AF':
			mo = 0
			ws[f'F{row + i}'] = f'{mon} | ({int(mo) + int(mon_wo)})'
		elif first(mon) >= 18:
			mo = (24 - first(mon)) + second(mon)
			ws[f'F{row + i}'] = f'{mon} | ({int(mo) + int(mon_wo)})'
		else:
			mo = (first(mon) - second(mon)) * -1
			ws[f'F{row + i}'] = f'{mon} | ({int(mo) + int(mon_wo)})'
		
		if 'cas' in tue.lower():
			tu = 0
			ws[f'G{row + i}'] = tue
		elif tue == 'AF':
			tu = 0
			ws[f'G{row + i}'] = f'{tue} | ({int(tu)})'
		elif first(tue) >= 18:
			tu = (24 - first(tue)) + second(tue)
			ws[f'G{row + i}'] = f'{tue} | ({int(tu)})'
		else:
			tu = (first(tue) - second(tue)) * -1
			ws[f'G{row + i}'] = f'{tue} | ({int(tu)})'

		if 'cas' in wed.lower():
			we = 0
			ws[f'H{row + i}'] = wed
		elif wed == 'AF':
			we = 0
			ws[f'H{row + i}'] = f'{wed} | ({int(we)})'
		elif first(wed) >= 18:
			we = (24 - first(wed)) + second(wed)
			ws[f'H{row + i}'] = f'{wed} | ({int(we)})'
		else:
			we = (first(wed) - second(wed)) * -1
			ws[f'H{row + i}'] = f'{wed} | ({int(we)})'

		# Add total hours at end
		hours = th + fr + sa + mon_wo + mo + tu + we
		ws[f'J{row + i}'] = f'{int(hours)}'

		sun_hours = int(su) + int(sun_wo)
		ws[f'K{row + i}'] = f'{int(sun_hours)}'

		# Add total hours to dic
		if week == 1:
			total_hours[name] = [int(hours), int(sun_hours)]
		else:
			total_hours[name][0] += int(hours) 
			total_hours[name][1] += int(sun_hours)

		if week == 2:
			ws[f'M{row + i}'] = name
			ws[f'N{row + i}'] = total_hours[name][0]
			ws[f'O{row + i}'] = total_hours[name][1]

		i += 1 
		
	# Close workbook
	wb.save("Total Time Worked.xlsx")
	wb.close()